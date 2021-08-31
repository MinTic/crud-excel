import random
from openpyxl import load_workbook
from datetime import datetime, timedelta
from colors import bcolors
from util import printWord, route

def read(route: str):
    file = load_workbook(route)
    data = file["data"]
    data = data["A2":"F256"]
    parsed = {}

    for i in data:
        if i[0].value == None:
            continue

        parsed.setdefault(i[0].value, {
            "task_name": i[1].value,
            "description": i[2].value,
            "status": i[3].value,
            "start_date": i[4].value,
            "end_date": i[5].value
        })

    return parsed


def addTask():
    file = load_workbook(route)
    data = file["data"]
    data = data["A2":"F" + str(data.max_row + 1)]
    hoja = file.active

    taskID = random.randint(1, 9999)
    taskName = str(input(bcolors.OKCYAN + "\nQue nombre le deseas poner a tu tarea?: " + bcolors.ENDC))
    while len(taskName) <= 0:
        taskName = str(input(bcolors.FAIL + "\nEl nombre de la tarea es muy corto!\nQue nombre le deseas poner a tu tarea?: " + bcolors.ENDC))

    taskDescription = str(input(bcolors.OKCYAN + "\nQue descripcion le deseas poner a tu tarea?: " + bcolors.ENDC))
    while len(taskDescription) <= 0:
        taskDescription = str(input(bcolors.FAIL + "\nLa descripcion de la tarea es muy corto!\nQue descripcion le deseas poner a tu tarea?: " + bcolors.ENDC))

    endDate = int(input(bcolors.OKCYAN + "\nCuantas horas tomara esta tarea?: "))
    while type(endDate) != int or endDate <= 0 or endDate >= 10000:
        endDate = int(input(bcolors.FAIL + "\nTiempo invalido!\nCuantas horas tomara esta tarea?: " + bcolors.ENDC))

    endDate = datetime.now() + timedelta(hours=endDate)
    startDate = f"{datetime.now().year}/{datetime.now().month}/{datetime.now().day} a las {datetime.now().hour} horas"

    for i in data:
        if not(isinstance(i[0].value, int)):
            id = i[0].row
            hoja.cell(row=id, column=1).value = taskID
            hoja.cell(row=id, column=2).value = taskName
            hoja.cell(row=id, column=3).value = taskDescription
            hoja.cell(row=id, column=4).value = "alive"
            hoja.cell(row=id, column=5).value = startDate
            hoja.cell(row=id, column=6).value = f"{endDate.year}/{endDate.month}/{endDate.day} a las {endDate.hour} horas"

    file.save(route)

    print(bcolors.BOLD + bcolors.OKGREEN + f"\nNueva tarea registrada!\n\nID: {taskID}\nNombre: {taskName}\nDescripcion: {taskDescription}\nEstado: alive \nInicio: {startDate}\nAcaba: {endDate.year}/{endDate.month}/{endDate.day} a las {endDate.hour} horas\n\n" + bcolors.ENDC)

def removeTask():
    file = load_workbook(route)
    data = file["data"]
    data = data["A2":"F" + str(data.max_row + 1)]
    hoja = file.active

    taskID = str(input(bcolors.OKCYAN + "\nQue ID tiene la tarea que desea eliminar?: " + bcolors.ENDC))
    while len(taskID) <= 0:
        taskID = str(input(bcolors.FAIL + "\nLa ID es muy corta!\nQue ID tiene la tarea que desea eliminar?: " + bcolors.ENDC))

    if taskID == "exit":
        return

    deleted = False
    for i in data:
        if str(i[0].value) == taskID:
            id = i[0].row
            hoja.cell(row=id, column=1).value = ""
            hoja.cell(row=id, column=2).value = ""
            hoja.cell(row=id, column=3).value = ""
            hoja.cell(row=id, column=4).value = ""
            hoja.cell(row=id, column=5).value = ""
            hoja.cell(row=id, column=6).value = ""
            deleted = True
            break

    if deleted is True:
        file.save(route)
        return print(bcolors.BOLD + bcolors.OKGREEN + f"\nTarea eliminada!\n" + bcolors.ENDC)

    print(bcolors.FAIL + "\nLa tarea con esta ID no existe!" + bcolors.ENDC)
    removeTask()

def updateTask():
    file = load_workbook(route)
    data = file["data"]
    data = data["A2":"F" + str(data.max_row + 1)]
    hoja = file.active

    taskID = str(input(bcolors.OKCYAN + "\nQue ID tiene la tarea que desea actualizar?: " + bcolors.ENDC))
    while len(taskID) <= 0:
        taskID = str(input(bcolors.FAIL + "\nLa ID es muy corta!\nQue ID tiene la tarea que desea actualizar?: " + bcolors.ENDC))

    print("1. Nombre")
    print("2. Descripcion")
    print("3. Estado")

    toUpdate = str(input(bcolors.OKCYAN + "Que dato quieres actualizar?: "  + bcolors.ENDC))
    while len(toUpdate) <= 0 or ["1", "2", "3"].count(toUpdate) <= 0:
        toUpdate = str(input(bcolors.FAIL + "El dato que quisistes actualizar no existe!\nQue dato quieres actualizar?: "  + bcolors.ENDC))

    newValue = str(input(bcolors.OKCYAN + "Escribe el valor que le quieres dar: "  + bcolors.ENDC))
    while len(newValue) <= 0:
        newValue = str(input(bcolors.FAIL + "El valor es invalido!\nEscribe el valor que le quieres dar: "  + bcolors.ENDC))

    updated = False
    for i in data:
        if str(i[0].value) == taskID:
            id = i[0].row
            if toUpdate == "1":
                hoja.cell(row=id, column=2).value = newValue
            elif toUpdate == "2":
                hoja.cell(row=id, column=3).value = newValue
            elif toUpdate == "3":
                hoja.cell(row=id, column=4).value = newValue
            else:
                break

            updated = True
        break
    
    if updated is True:
        file.save(route)
        print(bcolors.BOLD + bcolors.OKGREEN + f"\nTarea actualizada!\n" + bcolors.ENDC)

    print(bcolors.FAIL + "\nLa tarea con esta ID no existe!" + bcolors.ENDC)
    updateTask()

def show():
    list = ""
    data = read(route)
    for i in data:
        title = "\n" + bcolors.OKCYAN + bcolors.BOLD + f"Tarea {i}" + bcolors.ENDC
        list += title + f"""
        \nNombre: {data[i]["task_name"]}\nDescripcion: {data[i]["description"]}\nEstado: {data[i]["status"]}\nFecha inicio: {data[i]["start_date"]}\nFecha fin: {data[i]["end_date"]}
        """

    print("\n\n")
    printWord("Tareas")
    print(list)
